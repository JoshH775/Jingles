import PySimpleGUI as gui
import os, os.path
import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
from openpyxl.styles import Font
from tinytag import TinyTag
import Audio_Playback, YTD, playlistFunc
import datetime, random, time
global window
imagedir="images"
logo=imagedir+"/bigicon.ico"
dir_path = os.path.dirname(os.path.realpath(__file__))
os.chdir(dir_path)
playlistdir="playlists"

def sizeadjust(sizetuple=None,sizeint=None):
    print(sizetuple)
    if sizetuple!=None:
        width=(sizetuple[0]/1530)*100
        height=(sizetuple[1]/895)*100
        newtuple=(width,height)
        print(newtuple)
        return newtuple


def toPercentage(timeinfo):
    times=[]
    current=timeinfo[0]
    total=timeinfo[1]
    for i in (current,total):
        try:
            seconds=datetime.datetime.strptime(i,"%H:%M:%S")
        except:
            seconds=datetime.datetime.strptime(i,"%M:%S")
        a_timedelta = seconds - datetime.datetime(1900, 1, 1)
        seconds = a_timedelta.total_seconds()
        times.append(seconds)

    percentage=(times[0]/times[1])*100
    return percentage

def toSeconds(percentage,selectedsong):
    times = []
    timeinfo=Audio_Playback.playtime(selectedsong)
    current = timeinfo[0]
    total = timeinfo[1]
    for i in (current, total):
        try:
            seconds = datetime.datetime.strptime(i, "%H:%M:%S")
        except:
            seconds = datetime.datetime.strptime(i, "%M:%S")
        a_timedelta = seconds - datetime.datetime(1900, 1, 1)
        seconds = a_timedelta.total_seconds()
        times.append(seconds)

    seconds=times[1]*(percentage/100)
    return seconds


#TODO

#Convert FLACS and other files into mp3
#Create playlist function

#removing from folder
#13b6d6 - blue

logotheme = {'BACKGROUND': '#1f1f1f',
                'TEXT': 'white',
                'INPUT': '#ff7145',
                'TEXT_INPUT': '#000000',
                'SCROLL': '#480b9c',
                'BUTTON': ('white', '#480b9c'),
                'PROGRESS': ('#ff7145', '#3f3f3f'),
                'BORDER': 1,
                'SLIDER_DEPTH': 0,
                'PROGRESS_DEPTH': 0}

gui.theme_add_new("LogoTheme",logotheme)
gui.theme("LogoTheme")


class Spreadsheet():
    def libraryCreate(self):  # creates the library with headers
        bold = Font(bold=True, size=13, name="Arial")  # creating a bold font preset
        global sheet1, book
        book = Workbook()  # creating an instance of workbook
        sheet1 = book.active
        sheet1.title = "Library"
        x = 1
        if os.path.isfile("libpaths.txt") == False:
            f = open("libpaths.txt", "w+")
            f.close()
        columns = ["null", "Title", "Artists", "Album", "Year", "Genre", "Mood", "Path"]
        for i in range(1, len(columns)):  # for loop to create bold headings
            cell = sheet1.cell(row=x, column=i)
            cell.value = columns[i]
            cell.font = bold
        column_names = []
        for cell in sheet1[1]:
            column_names.append(cell.value)

        book.save("library - do not delete.xlsx")
        return sheet1

    def findEmpty(self,column="A"):
        row = 1
        foundempty = False
        while foundempty == False:
            cell = column + str(row)
            if sheet1[cell].value == None:
                foundempty = True
            else:
                row += 1
        firstempty = row
        return firstempty

    def tags(self, totaglist):
        firstempty = self.findEmpty()
        for j in range(len(totaglist)):
            firstempty = str(firstempty)
            tagged = totaglist[j]
            tagged = TinyTag.get(tagged)
            sheet1["A" + firstempty] = tagged.title
            sheet1["B" + firstempty] = tagged.artist
            sheet1["C" + firstempty] = tagged.album
            sheet1["D" + firstempty] = tagged.year
            sheet1["E" + firstempty] = tagged.genre
            firstempty = int(firstempty) + 1
            print(totaglist[j])

    def addFolder(self, book):
        try:

            dir = filedialog.askdirectory()


            f = open("libpaths.txt", "r+")
            # if dir in f.read():
            #     gui.PopupError("Cannot add folder, folder already in library")
            # elif dir not in f.read():
            f.write(dir + "\n")
            global songpaths, libpaths
            libpaths = f.readlines()
            f.close()
            songpaths = []
            for file in os.scandir(dir):
                print(file)
                if file.name.lower().endswith(".mp3"):# }
                    songpaths.append(file.path)  # adding all songs in dir to list
            songpaths.insert(0, "null")  # }

            firstempty=self.findEmpty(column="G")-1
            print(firstempty,"firstempty")
            for i in range(0, len(songpaths)):
                pathcell = "G" + (str(firstempty+i))
                print(pathcell)
                if songpaths[i] == "null":  # catch so null is not included
                    continue
                else:
                    sheet1[pathcell] = songpaths[i]


            songpaths.pop(0)
            self.tags(songpaths)
            book.save("library - do not delete.xlsx")


            print("added songsw")
        except FileNotFoundError:
            gui.Popup("SOmething went wrong")
    def updateSpreadsheet(self,toadd,sheet,book):
        max=str(self.findEmpty())
        path=toadd
        toadd=TinyTag.get(toadd)
        sheet["A" + max].value=toadd.title
        sheet["B" + max].value = toadd.artist
        sheet["C" + max].value = toadd.album
        sheet["D" + max].value = toadd.year
        sheet["E" + max].value = toadd.genre
        sheet["G" + max].value = path
        book.save("library - do not delete.xlsx")




def LBox(key, scrollbar, size=(25, 100)):  # function for defining listboxes on homescreen
    rclickmenu = ["Play",["Add to Queue","Add to Playlist"]]
    if size == (25, 100):
        lboxlayout = [[gui.Listbox(values=[], enable_events=True, select_mode=gui.LISTBOX_SELECT_MODE_SINGLE,
                                   pad=(0, 0), no_scrollbar=scrollbar, key=key, right_click_menu=rclickmenu,
                                   size=(25, 100))]]
    return gui.Column(layout=lboxlayout, scrollable=False)

def getArtists(sheet,mode=1): #mode1 - returning all artists.
    if mode==1:
        artists=[]
        for row in sheet.iter_rows():
            if row[0].value in ("Title", "Artist", "Album", "Year", "Genre", "Mood", "Path"):
                continue
            else:
                artists.append(row[1].value)
        artists = list(dict.fromkeys(artists)) #removing duplicates as dictionaries cant have dupes
        if None in artists:
            artists.remove(None)

        return sorted(artists)
def getAlbums(sheet):
    albums=[]
    for row in sheet.iter_rows():
        if row[0].value in ("Title", "Artist", "Album", "Year", "Genre", "Mood", "Path"):
            continue
        else:
            albums.append(row[2].value)
    albums = list(dict.fromkeys(albums))  # removing duplicates as dictionaries cant have dupes
    if None in albums:
        albums.remove(None)

    return sorted(albums)

def getPlaylists():
    playlists=os.listdir(playlistdir)
    if len(playlists)>0:
        for i in range(len(playlists)):
            playlists[i]=playlists[i][:-4]
    return playlists

def controlFunction(sliderkey="-SONGSLIDER-",progresskey="-PROGRESSBAR-",pausekey="-PAUSE-",prevkey="-PREVSONG-",nextkey="-NEXTSONG-",exitkey="-EXIT-",columnkey="-CONTROLCOLUMN-"):
    controlcolumn = [[gui.Slider(enable_events=True, tooltip="Song Progress", disable_number_display=False,
                                 key=sliderkey, size=(37, 25), range=(0, 100)),
                      gui.ProgressBar(orientation="vertical", max_value=100, size=(65, 25), key=progresskey)],
                     [gui.Button("<<", key=prevkey,size=(3,2)),gui.Button("Pause", key=pausekey,size=(5,2)),gui.Button(">>", key=nextkey,size=(3,2))],
                     [gui.Button("Exit", key=exitkey,size=(10,1),button_color=("white","red"))]]
    return gui.Column(layout=controlcolumn,element_justification="c",key=columnkey,pad=(0, 0))

class Layouts():
    # constructor
    sortbybar = [
        [gui.Text("Sort By:"), gui.Button("Title", key="-TITLESORT-"), gui.Button("Artist", key="-ARTISTSORT-"),
         gui.Button("Album", key="-ALBUMSORT-"), gui.Radio(text="Ascending", default=True, group_id="ORDER",enable_events=True,key="-ASCRADIO-"),
         gui.Radio(text="Decending", group_id="ORDER",enable_events=True,key="-DSCRADIO-"), gui.Text("Volume:"),
         gui.Slider(range=(1, 100), default_value=50, orientation="horizontal",key="-VOLSLIDE-",enable_events=True),
         gui.Button(button_text="SHUFFLE: OFF",size=(12,2),key="-SHUFFLE-"),
         gui.Button(button_text="REPEAT: OFF",size=(12,2),key="-REPEAT-"),
         gui.Text("Time elapsed:"),gui.Text("00:00",key="-PLAYHEAD-",pad=(0,0)),gui.Text("of",pad=(0,0)),gui.Text("00:00",key="-LENGTH-",pad=(0,0)),
         gui.Text("   Now playing:",pad=(0,0)),gui.InputText(default_text="----",key="-NOWPLAYING-",readonly=True,size=(37,1))]]

    sortingtabcolumn = [
        [gui.Button(button_text="Library Settings", key="-LIBRARYSETTINGS-", button_color=("white", "gray"),
                    size=(17, 1))],
        [gui.Text("______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("SONGS", key="-SONGBUTTON-", font=("Integral 20 bold"), size=(9, 2))],
        [gui.Text("______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("ARTISTS", key="-ARTISTBUTTON-", font=("Integral 20 bold"), size=(9, 2))],
        [gui.Text("_______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("ALBUMS", key="-ALBUMBUTTON-", font=("Integral 20 bold"), size=(9, 2))],
        [gui.Text("______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("GENRES",key="-GENREBUTTON-",font=("Integral 20 bold"), size=(9, 2))],
        [gui.Text("______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("PLAYLISTS", key="-PLAYLISTBUTTON-", font=("Integral 20 bold"), size=(9, 2))],
        [gui.Text("______________________")],
        [gui.Text(" ", font="any 1")],

        [gui.Button("YOUTUBE", key="-YTDBUTTON-", font=("Integral 20 bold"), size=(9, 2))]]



    songdisplaycolumn = [[LBox("-TITLEBOX-", True), LBox("-ARTISTBOX-", True), LBox("-ALBUMBOX-", True), LBox("-YEARBOX-", True),LBox("-GENREBOX-", True), LBox("-PATHBOX-", True)]]

    artistdisplaycolumn=[[gui.Listbox(values=[],enable_events=True,key="-ARTISTDISPLAYBOX-",size=(159,100))]]
    albumdisplaycolumn=[[gui.Listbox(values=[],enable_events=True,key="-ALBUMDISPLAYBOX-",size=(159,100))]]
    playlistdisplaycolumn=[[gui.Listbox(values=[],enable_events=True,key="-PLAYLISTDISPLAYBOX-",size=(159,100))]]
    genredisplaycolumn=[[gui.Listbox(values=[],enable_events=True,key="-GENREDISPLAYBOX-",size=(159,100))]]




    librarycolumn1 = [[gui.Button("Home", key="-HOME-", button_color=("white", "gray")),
                       gui.Text(text="Library Settings", text_color="white", font="Helvetica 26 bold underline",
                                justification="center")],
                      [gui.Button("Reset Library", key="-RESETLIB-", size=(17, 1), button_color=("black", "red"))],
                      [gui.Button("Add folder", key="-ADDFOLDER-", button_color=("white", "gray"), size=(17, 1))]]

    # entire screen layouts
    homescreenLL = [[gui.HSeparator()],
                  [gui.Column(sortingtabcolumn), gui.VSeparator(),gui.Column(songdisplaycolumn, visible=True, key="-SONGDISPLAYCOLUMN-", vertical_alignment="top",scrollable=True, vertical_scroll_only=True),gui.Column(artistdisplaycolumn,visible=False,key="-ARTISTDISPLAYCOLUMN-",vertical_alignment="top",scrollable=False),gui.Column(albumdisplaycolumn,visible=False,key="-ALBUMDISPLAYCOLUMN-",vertical_alignment="top",scrollable=False),gui.Column(playlistdisplaycolumn,visible=False,key="-PLAYLISTDISPLAYCOLUMN-",vertical_alignment="top",scrollable=False),gui.Column(genredisplaycolumn,key="-GENREDISPLAYCOLUMN-",visible=False,vertical_alignment="top",scrollable=False)]]


    homescreen=[[gui.Column(layout=sortbybar)],
                [gui.Column(homescreenLL,scrollable=False,element_justification="l"),controlFunction()]]
    librarysettings = [[gui.Column(librarycolumn1)],
                       [gui.Text("Folders in library:")],
                       [gui.Listbox(values=[], key="-LIBPATHS-", size=(0, 50), background_color="white",
                                    text_color="black", enable_events=True)]]


# Funciton to create a library.xlxs file if one is not present
def noLibrary():
    noliblayout = [gui.Text("No library detected. Create one?")], \
                  [gui.Button("Create", key="-CREATELIB-")]

    noLibrary = gui.Window(title="No Library Detected", layout=noliblayout)
    while True:
        event, values = noLibrary.read()
        if event == "-CREATELIB-":
            ss.libraryCreate()
            gui.Popup("Library Created!")
            noLibrary.close()
        if event == gui.WIN_CLOSED:
            noLibrary.close()
            break

def getGenres(sheet):
    genres=[]
    for row in sheet.iter_rows():
        if row[0].value in ("Title", "Artist", "Album", "Year", "Genre", "Mood", "Path","Genres"):
            continue
        else:
            genres.append(row[4].value)
    genres = list(dict.fromkeys(genres))  # removing duplicates as dictionaries cant have dupes
    if None in genres:
        genres.remove(None)
    if "Genre" in genres:
        genres.remove("Genre")

    return sorted(genres)



# main function to start Application
def musicDict(sheet, window, sortmethod="-TITLESORT-",mode=1,artist=None,album=None,genre=None,order=True,playlist=None):  # reads songs from xlxs, sorts songs. Default sort is alphabetically by title. Also changes listbox and column size every call. Mode 1 - standard whole library display. Mode 2 - specific artist sort, Mode 3 - specific album sort, Mode 4 - Playlist display
    global titles, artists, albums, years, genres, moods, paths, zippedmetadata
    titles = []
    artists = []
    albums = []
    years = []
    genres = []
    moods=[]
    paths = []
    max=Spreadsheet.findEmpty(Spreadsheet)
    max=max-1
    if order==True:
        order="asc"
    else:
        order="dsc"

    if mode==4:
        for i in range(len(playlist)):
            item = playlist[i]
            titles.append(item[0])
            artists.append(item[1])
            albums.append(item[2])
            years.append(item[3])
            genres.append(item[4])
            moods.append("None")
            paths.append(item[5])
    else:
        for row in sheet.iter_rows(1,max):
            if row[0].value in ("Title", "Artist", "Album", "Year", "Genre", "Mood", "Path"):
                continue
            else:
                if mode==2:
                    print(row[1].value)
                    if row[1].value == artist:
                        titles.append(row[0].value)
                        artists.append(row[1].value)
                        albums.append(row[2].value)
                        years.append(row[3].value)
                        genres.append(row[4].value)
                        moods.append("None")
                        paths.append(row[6].value)
                        continue
                    else:
                        continue
                elif mode==3:
                    if row[2].value == album:
                        titles.append(row[0].value)
                        artists.append(row[1].value)
                        albums.append(row[2].value)
                        years.append(row[3].value)
                        genres.append(row[4].value)
                        moods.append("None")
                        paths.append(row[6].value)
                        continue
                    else:
                        continue
                elif mode==5:
                    if row[4].value == genre:
                        titles.append(row[0].value)
                        artists.append(row[1].value)
                        albums.append(row[2].value)
                        years.append(row[3].value)
                        genres.append(row[4].value)
                        moods.append("None")
                        paths.append(row[6].value)
                        continue
                    else:
                        continue


                else:
                    titles.append(row[0].value)
                    artists.append(row[1].value)
                    albums.append(row[2].value)
                    years.append(row[3].value)
                    genres.append(row[4].value)
                    moods.append("None")
                    paths.append(row[6].value)




    zippedmetadata = zip(titles, artists, albums, years, genres, moods, paths)
    # if statements deciding how songs will be sorted
    if sortmethod == "-TITLESORT-":
        zippedmetadata = sorted(zip(titles, artists, albums, years, genres, moods, paths))  # alphabetical sort by title


    elif sortmethod == "-ARTISTSORT-":
        zippedmetadata = sorted(
            zip(artists, titles, albums, years, genres, moods, paths))  # alphabetical sory by artist
    elif sortmethod == "-ALBUMSORT-":
        zippedmetadata = sorted(zip(albums, titles, artists, years, genres, moods, paths))

    for lis in (titles, artists, albums, years, genres, moods, paths):  # clearing lists for sorting
        lis.clear()

    for i in range(len(zippedmetadata)):  # Appending sorted values to list after sort
        currentsongdata = zippedmetadata[i]
        titles.append(currentsongdata[0])
        artists.append(currentsongdata[1])
        albums.append(currentsongdata[2])
        years.append(currentsongdata[3])
        genres.append(currentsongdata[4])
        moods.append(currentsongdata[5])
        paths.append(currentsongdata[6])

    if order == "dsc":
        for lis in (titles, artists, albums, years, genres, moods, paths):
            lis.reverse()

    window["-TITLEBOX-"].update(values=titles)

    window["-ARTISTBOX-"].update(values=artists)

    window["-ALBUMBOX-"].update(values=albums)

    window["-YEARBOX-"].update(values=years)

    window["-GENREBOX-"].update(values=genres)

    window["-PATHBOX-"].update(values=paths)

    for event in ("-TITLEBOX-", "-ARTISTBOX-", "-ALBUMBOX-", "-YEARBOX-", "-GENREBOX-", "-PATHBOX-"): #setting listbox sizes to fit number of items
        sizes = (25, len(titles))
        window[event].set_size(sizes)



    pathsize = (50, len(titles))
    window["-PATHBOX-"].set_size(pathsize)
    listboxpixels = window["-TITLEBOX-"].get_size() #getting size of one listbox
    # print(listboxpixels)
    # listboxpixels=(listboxpixels[0],200)
    # print(listboxpixels)
    # print(listboxpixels,"newlistboxpixels")

    if window["-SONGDISPLAYCOLUMN-"].get_size() != listboxpixels:
        window["-SONGDISPLAYCOLUMN-"].set_size(listboxpixels)  # setting column size to listbox size
        window["-SONGDISPLAYCOLUMN-"].Widget.canvas.yview_moveto(0.0)  # moving scroll heading to top
    window["-SONGDISPLAYCOLUMN-"].set_size(listboxpixels)  # setting column size to listbox size


def nowPlaying(selectedsong,sortmode,window): #function for updating now playing box at top right

    if sortmode == "-TITLESORT-":
        np=(selectedsong[0]+" - "+selectedsong[1])
    elif sortmode == "-ARTISTSORT-":
        np=(selectedsong[1]+" - "+selectedsong[0])

    window["-NOWPLAYING-"].update(np)


def enQueue(selectedsong,window):
    Fqueue=[]
    leng=len(window["-TITLEBOX-"].get_list_values())
    count=0
    for j in range(leng):
        song=[]
        for i in ("-TITLEBOX-", "-ARTISTBOX-", "-ALBUMBOX-", "-YEARBOX-", "-GENREBOX-", "-PATHBOX-"):
            song.append(window[i].get_list_values()[count])
        count+=1

        Fqueue.append(song)

    #search
    counter=1
    for i in range(len(Fqueue)):
        if Fqueue[i] == selectedsong:
            break
        else:
            counter+=1

    prevqueue=Fqueue[:counter]

    for item in prevqueue:
        if item in Fqueue:
            Fqueue.remove(item)
    prevqueue.remove(selectedsong)

    return Fqueue,prevqueue


def play(selectedsong,vol,start=0.0):
    dontplay=False
    try:
        Audio_Playback.play(selectedsong,vol,start)
    except:
        dontplay=True


    return dontplay


def main():
    global ss, ll
    ss = Spreadsheet()  # defining classes
    ll = Layouts()
    global book, sheet1, sortmode, queue, prevqueue
    sortmode="-TITLESORT-"
    if os.path.isfile("library - do not delete.xlsx") == False:
        noLibrary()
        print("Creating Library")
    else:
        book = openpyxl.load_workbook("library - do not delete.xlsx")
        sheet1 = book.active

    if os.path.exists(os.path.join(os.getcwd(),playlistdir)) == True:
        pass
    else:
        os.makedirs(os.path.join(os.getcwd(),playlistdir))
    # layouts: HOME - Homescreen, LIBSETLAYOUT - Library Settings

    layout = [
        [
            gui.Column(Layouts.homescreen, key="HOMECOLUMN", visible=True),

            gui.Column(Layouts.librarysettings, key="LIBSETCOLUMN", visible=False, element_justification="l",
                       size=(600, 490))

        ]
    ]

    #size=(1530, 895)
    window = gui.Window(title="Josh's Jingles", layout=layout, size=(1530, 895),icon=logo,resizable=True).Finalize()
    musicDict(sheet1, window, sortmethod=sortmode, order=True)
    f = open("libpaths.txt", "r")
    window["-LIBPATHS-"].update(values=f.readlines())
    f.close()
    queue=[]
    prevqueue=[]
    nextsong=False
    prevsong=False
    repeatmode=0
    shuffle=False
    newplaytime=0.0
    while True:
        if Audio_Playback.getBusy()==1:
            event, values = window.read(timeout=10)
            if dontplay==False:
                times = Audio_Playback.playtime(selectedsong,newplaytime)

                window["-PROGRESSBAR-"].update(toPercentage(times))
                window["-PLAYHEAD-"].update(str(times[0])) #updating the status bar
                window["-LENGTH-"].update(str(times[1]))

                songover=Audio_Playback.endCheck()
                if songover==True or nextsong==True: #AT THE END OF THE SONG
                    newplaytime=0.0
                    prevqueue.append(selectedsong)
                    if repeatmode==2:
                        selectedsong=prevqueue.pop(-1)
                    else:
                        try:
                            if shuffle==True:
                                print(shuffle)
                                selectedsong=random.choice(queue)
                                queue.remove(selectedsong)
                            else:

                                    selectedsong=queue.pop(0)
                        except IndexError:
                            print("end of queue",repeatmode)
                            if repeatmode == 1:
                                queue.clear()
                                selectedsong=prevqueue[0]
                                queue=enQueue(selectedsong,window)[0]

                    nowPlaying(selectedsong, sortmode, window)
                    dontplay=play(selectedsong,vol=values["-VOLSLIDE-"]/100)

                    if dontplay==False:
                        times = Audio_Playback.playtime(selectedsong, 0.0)

                        window["-SONGSLIDER-"].update(0)
                        window["-PROGRESSBAR-"].update(toPercentage(times))
                        window["-PLAYHEAD-"].update(str(times[0]))
                        window["-LENGTH-"].update(str(times[1]))

                        nextsong=False



                if prevsong==True:


                    prevsong=False
                    newplaytime=0.0
                    queue.insert(0,selectedsong)
                    try:
                        selectedsong=prevqueue.pop()
                    except IndexError:
                        print("queue empty")


                    dontplay=play(selectedsong,vol=values["-VOLSLIDE-"]/100)

                    if dontplay==False:
                        times = Audio_Playback.playtime(selectedsong, 0.0)
                        nowPlaying(selectedsong, sortmode, window)
                        window["-SONGSLIDER-"].update(0)
                        window["-PROGRESSBAR-"].update(toPercentage(times))
                        window["-PLAYHEAD-"].update(str(times[0]))
                        window["-LENGTH-"].update(str(times[1]))






        else:
            event, values = window.read()
        if event != "__TIMEOUT__":
            print(event)
        if event in (gui.WIN_CLOSED, "-EXIT-"):
            window.close()
            break

        if event == "-LIBRARYSETTINGS-":
            window["LIBSETCOLUMN"].update(visible=True)
            window["HOMECOLUMN"].update(visible=False)

        if event == "-HOME-":
            window["LIBSETCOLUMN"].update(visible=False)
            window["HOMECOLUMN"].update(visible=True)
            musicDict(sheet1,window,sortmode,mode=1,order=values["-ASCRADIO-"])





        if event == "-CREATE-":
            ss.libraryCreate()
            gui.popup("Library created!")

        if event == "-ADDFOLDER-":
            f = open("libpaths.txt", "r")
            ss.addFolder(book)
            libpaths = f.readlines()
            f.close()
            window["-LIBPATHS-"].update(values=libpaths)

        if event == "-RESETLIB-":
            okcancel = gui.PopupOKCancel(
                "Resetting the library will remove all attached song folders, and the program will need to be restarted.",
                "Continue?")
            if okcancel == "OK":
                window.close()
                os.remove("libpaths.txt")
                os.remove("library - do not delete.xlsx")

        if event == "-LIBPATHS-":
            selected = True
            selectedpath = str(values["-LIBPATHS-"])
            for i in ["[", "]", "'"]:
                selectedpath = selectedpath.replace(i, '')



        if event in ("-TITLEBOX-", "-ARTISTBOX-", "-ALBUMBOX-", "-YEARBOX-", "-GENREBOX-", "-PATHBOX-"): #IF SONG SELECTED
            indexes = window[event].GetIndexes()
            selectedsong = []
            for event in ("-TITLEBOX-", "-ARTISTBOX-", "-ALBUMBOX-", "-YEARBOX-", "-GENREBOX-", "-PATHBOX-"): #setting other listboxes to match selected value
                window[event].Update(set_to_index=indexes)
                try:
                    prevqueue=[]
                    newplaytime=0.0
                    window["-SONGSLIDER-"].update(newplaytime)
                    selectedsong.append(list(window[event].get_list_values())[indexes[0]])

                    originsong=selectedsong

                    dontplay=play(selectedsong, vol=(values["-VOLSLIDE-"] / 100)) #plays the song
                    print(dontplay)
                    if dontplay==False:

                        window[event].set_value([])
                        nowPlaying(selectedsong, sortmode,
                                   window)  # updates now playing tag with the song, and sortmode for order of tags
                        queue=enQueue(selectedsong,window)[0]
                        prevqueue=enQueue(selectedsong,window)[1]




                # ENTIRE METADATA OF SELECTED SONG
                except IndexError:  # stopping an error being thrown when a empty space is clicked
                    print("passed")
                    pass







        if event in ("-TITLESORT-", "-ARTISTSORT-", "-ALBUMSORT-"): # top 3 buttons following sort by:

            musicDict(sheet1, window, sortmethod=event, order=values["-ASCRADIO-"]) #updating with new sortmode
            sortmode=event

        if event in ("-SONGBUTTON-", "-ARTISTBUTTON-","-ALBUMBUTTON-","-PLAYLISTBUTTON-","-GENREBUTTON-"):
            if event == "-SONGBUTTON-":
                window["-ARTISTDISPLAYCOLUMN-"].update(visible=False)
                window["-ALBUMDISPLAYCOLUMN-"].update(visible=False)
                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=False)
                window["-GENREDISPLAYCOLUMN-"].update(visible=False)
                window["-SONGDISPLAYCOLUMN-"].update(visible=True)
                musicDict(sheet1,window,sortmode,mode=1,order=values["-ASCRADIO-"])


            if event=="-ARTISTBUTTON-":
                window["-ARTISTDISPLAYBOX-"].update(getArtists(sheet1,mode=1))
                window["-ARTISTDISPLAYCOLUMN-"].update(visible=True)
                window["-SONGDISPLAYCOLUMN-"].update(visible=False)
                window["-ALBUMDISPLAYCOLUMN-"].update(visible=False)
                window["-GENREDISPLAYCOLUMN-"].update(visible=False)
                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=False)

            if event=="-ALBUMBUTTON-":
                window["-ALBUMDISPLAYBOX-"].update(getAlbums(sheet1))

                window["-ALBUMDISPLAYCOLUMN-"].update(visible=True)
                window["-SONGDISPLAYCOLUMN-"].update(visible=False)
                window["-ARTISTDISPLAYCOLUMN-"].update(visible=False)
                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=False)
                window["-GENREDISPLAYCOLUMN-"].update(visible=False)

            if event=="-PLAYLISTBUTTON-":
                window["-PLAYLISTDISPLAYBOX-"].update(getPlaylists())
                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=True)
                window["-SONGDISPLAYCOLUMN-"].update(visible=False)
                window["-ARTISTDISPLAYCOLUMN-"].update(visible=False)
                window["-ALBUMDISPLAYCOLUMN-"].update(visible=False)
                window["-GENREDISPLAYCOLUMN-"].update(visible=False)

            if event =="-GENREBUTTON-":
                window["-GENREDISPLAYBOX-"].update(getGenres(sheet1))
                window["-GENREDISPLAYCOLUMN-"].update(visible=True)
                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=False)
                window["-SONGDISPLAYCOLUMN-"].update(visible=False)
                window["-ARTISTDISPLAYCOLUMN-"].update(visible=False)
                window["-ALBUMDISPLAYCOLUMN-"].update(visible=False)



        if event=="-ARTISTDISPLAYBOX-":
            artist=values["-ARTISTDISPLAYBOX-"][0]

            window["-ARTISTDISPLAYCOLUMN-"].update(visible=False)
            window["-SONGDISPLAYCOLUMN-"].update(visible=True)
            musicDict(sheet1, window, sortmode, mode=2, artist=artist,order=values["-ASCRADIO-"])

        if event=="-ALBUMDISPLAYBOX-":

            album=values["-ALBUMDISPLAYBOX-"][0]


            window["-ALBUMDISPLAYCOLUMN-"].update(visible=False)
            window["-SONGDISPLAYCOLUMN-"].update(visible=True)
            musicDict(sheet1,window,sortmode,mode=3,album=album,order=values["-ASCRADIO-"])

        if event=="-GENREDISPLAYBOX-":
            genre=values[event][0]
            window["-GENREDISPLAYCOLUMN-"].update(visible=False)
            window["-SONGDISPLAYCOLUMN-"].update(visible=True)
            musicDict(sheet1,window,sortmode,mode=5,genre=genre,order=values["-ASCRADIO-"])

        if event=="-PLAYLISTDISPLAYBOX-":
            try:
                playlistlocation=values["-PLAYLISTDISPLAYBOX-"][0]+".txt"
                playlistlocation=playlistdir+"\{}".format(playlistlocation)
                print(playlistlocation)

                with open(playlistlocation) as file:
                    playlist = []
                    for line in file:
                        if line=="\n":
                            continue
                        else:
                            line=line.rstrip("\n")
                            playlistentry=line.split("***")
                            playlist.append(playlistentry)
                            print(playlistentry)


                window["-PLAYLISTDISPLAYCOLUMN-"].update(visible=False)
                window["-SONGDISPLAYCOLUMN-"].update(visible=True)
                musicDict(sheet1,window,sortmode,mode=4,playlist=playlist,order=values["-ASCRADIO-"])
            except:
                pass
        if event == "-VOLSLIDE-":
            Audio_Playback.setVol(values["-VOLSLIDE-"]/100)

        if event == "-PAUSE-":
            Audio_Playback.pause()
            if Audio_Playback.getBusy()==1:
                window["-PAUSE-"].update(button_color=("white","#480b9c"))
            else:
                window["-PAUSE-"].update(button_color=(("white","#ff7145")))
        if event == "-SONGSLIDER-":
            Audio_Playback.pause()
            window["-PROGRESSBAR-"].update(values["-SONGSLIDER-"])
            try:
                newplaytime=toSeconds(values["-SONGSLIDER-"],selectedsong[5])
                play(selectedsong,vol=(values["-VOLSLIDE-"]/100),start=newplaytime)
                Audio_Playback.playtime(selectedsong,newplaytime)
            except UnboundLocalError:
                print("no playing soing")


        if event == "-NEXTSONG-":
            nextsong=True
        if event == "-PREVSONG-":
            prevsong=True

        if event=="-SHUFFLE-":

            if shuffle==False:

                shuffle=True
                window["-SHUFFLE-"].update(text="SHUFFLE: ON")
                window["-SHUFFLE-"].update(button_color=("white","#ff7145"))


            elif shuffle==True:
                shuffle=False
                window["-SHUFFLE-"].update(text="SHUFFLE: OFF")
                window["-SHUFFLE-"].update(button_color=("white", "#480b9c"))

        if event=="-YTDBUTTON-":
            try:
                mp3=YTD.main()
                print(mp3)
                ss.updateSpreadsheet(toadd=mp3,sheet=sheet1,book=book)
                musicDict(sheet1, window,sortmethod=sortmode,order=values["-ASCRADIO-"])
            except:
                continue


        if event == "-REPEAT-":
            if repeatmode == 0:
                window["-REPEAT-"].update(text="REPEAT: QUEUE")
                window["-REPEAT-"].update(button_color=("white", "#ff7145"))
                repeatmode=1
            elif repeatmode == 1:
                window["-REPEAT-"].update(text="REPEAT: SONG")
                repeatmode=2
            elif repeatmode == 2:
                window["-REPEAT-"].update(text="REPEAT: OFF")
                window["-REPEAT-"].update(button_color=("white", "#480b9c"))
                repeatmode=0

        if event in ("-ASCRADIO-","-DSCRADIO-"): #if the two "Ascending" or "Descending" radios are interacted with
            musicDict(sheet1,window,sortmode,order=values["-ASCRADIO-"])

        if event=="Add to Playlist":
            try:
                playlistFunc.main(originsong,sortmode)
            except:
                pass

main()
